from openpyxl import Workbook, load_workbook


def find_first_empty_row(sheet, column):
    row = 1
    while sheet.cell(row=row, column=column).value is not None:
        row += 1
    return row


def add_to_excel(file_path, data):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active

    start_row = find_first_empty_row(sheet, 1)
    start_col = 1

    for idx, item in enumerate(data, start=start_col):
        sheet.cell(row=start_row, column=idx, value=item)

    workbook.save(file_path)
